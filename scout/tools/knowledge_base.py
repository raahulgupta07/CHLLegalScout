"""
Knowledge Base Module
====================

Handles uploading and processing Excel, CSV, and Word files
to create a knowledge base for the agent.
"""

import csv
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from docx import Document
from psycopg import connect
from scout.tools.template_analyzer import get_db_connection
from os import getenv


def analyze_excel_with_ai(file_path: str, filename: str) -> Dict[str, Any]:
    """
    AI-powered Excel analyzer that understands complex Excel structures.
    Handles merged cells, multi-line headers, and complex data.
    """
    try:
        from openai import OpenAI

        print(f"[AI ANALYZER] Starting analysis for {filename}")

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        raw_data = []
        max_cols = 0

        for row in sheet.iter_rows(max_row=250, values_only=True):
            if row and any(cell is not None for cell in row):
                row_list = list(row)
                max_cols = max(max_cols, len(row_list))
                raw_data.append([str(cell) if cell is not None else "" for cell in row_list])

        wb.close()

        if not raw_data:
            return {"success": False, "error": "No data found in Excel file"}

        print(f"[AI ANALYZER] Read {len(raw_data)} rows, {max_cols} columns")

        api_key = getenv("OPENROUTER_API_KEY") or getenv("OPENAI_API_KEY")
        if not api_key:
            print("[AI ANALYZER] No API key found, using legacy parser")
            return {"success": False, "error": "No AI API key configured"}

        # Skip header rows and get only data
        # Row 0: note, Row 1: headers, Row 2: empty
        data_rows = [row for row in raw_data[3:] if row and any(cell for cell in row if cell)]

        # Take first 50 rows to process
        batch = data_rows[:50]

        from app.model_config import OPENROUTER_BASE_URL, get_model
        client = OpenAI(
            api_key=api_key,
            base_url=OPENROUTER_BASE_URL,
            timeout=60.0,
        )

        prompt = f"""Convert this Excel company data to clean JSON array.

Headers: {json.dumps(raw_data[1][:15])}
Data rows (first 50):
{json.dumps(batch, indent=2)}

Rules:
- Row 2 has actual headers: Number, Company name, Date of Incorporation, etc.
- Extract: number, company_name (extract name from "Name (Reg No: XXX)"), date_of_incorporation, registered_office_address, principal_place_of_business, company_size, foreign_company, under_management, group_company, directors (take first 3), shareholders (take first 3), currency
- Skip empty rows
- Return JSON array with all {len(batch)} records. Be concise.

Return ONLY JSON array."""

        response = client.chat.completions.create(
            model=get_model("chat"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=8000,
        )

        content = response.choices[0].message.content
        print(f"[AI ANALYZER] Raw response length: {len(content)}")

        if content:
            content = content.strip()
            if content.startswith("```"):
                content = content.split("```")[1]
                if content.startswith("json"):
                    content = content[4:]
            content = content.strip()

        cleaned_data = json.loads(content)

        return {"success": True, "cleaned_data": cleaned_data, "record_count": len(cleaned_data)}

    except Exception as e:
        return {"success": False, "error": str(e)}


def clean_header(header: str) -> str:
    """Clean a header string - remove special chars, normalize spaces."""
    header = re.sub(r"\[.*?\]", "", header)
    header = re.sub(r" Remarks:.*", "", header, flags=re.IGNORECASE)
    header = re.sub(r" last updated.*", "", header, flags=re.IGNORECASE)
    header = re.sub(r"\s+", " ", header)
    header = header.strip()
    return header


def process_excel(file_path: str, filename: str, use_ai: bool = True) -> Dict[str, Any]:
    """Process Excel file - with AI analysis for complex files."""

    if use_ai:
        ai_result = analyze_excel_with_ai(file_path, filename)
        if ai_result.get("success") and ai_result.get("cleaned_data"):
            return store_cleaned_data(filename, ai_result["cleaned_data"], "excel")

    return process_excel_legacy(file_path, filename)


def process_file(file_path: str, filename: str) -> Dict[str, Any]:
    """Process uploaded file based on type."""
    file_ext = Path(filename).suffix.lower()

    if file_ext in [".xlsx", ".xls"]:
        return process_excel(file_path, filename)
    elif file_ext == ".csv":
        return process_csv(file_path, filename)
    elif file_ext == ".docx":
        return process_word(file_path, filename)
    else:
        return {"success": False, "error": f"Unsupported file type: {file_ext}"}


def store_cleaned_data(filename: str, cleaned_data: List[Dict], file_type: str = "excel") -> Dict[str, Any]:
    """Store AI-cleaned data into the database."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM knowledge_sources WHERE filename = %s", (filename,))
        cur.execute("DELETE FROM knowledge_raw WHERE source_file = %s", (filename,))
        cur.execute("DELETE FROM knowledge_lookup WHERE source_file = %s", (filename,))
        cur.execute("DELETE FROM knowledge_vec WHERE source_file = %s", (filename,))

        for row_idx, record in enumerate(cleaned_data, start=1):
            cur.execute(
                """INSERT INTO knowledge_raw 
                (source_file, file_type, sheet_name, row_number, data)
                VALUES (%s, %s, %s, %s, %s)""",
                (filename, file_type, "Sheet1", row_idx, json.dumps(record)),
            )

            for key, value in record.items():
                if value:
                    cur.execute(
                        """INSERT INTO knowledge_lookup (key_name, key_value, source_file)
                        VALUES (%s, %s, %s)""",
                        (key.lower(), str(value), filename),
                    )

            text_chunk = " | ".join([f"{k}: {v}" for k, v in record.items() if v])
            cur.execute(
                """INSERT INTO knowledge_vec (content, source_file, metadata)
                VALUES (%s, %s, %s)""",
                (text_chunk, filename, json.dumps(record)),
            )

        cur.execute(
            """INSERT INTO knowledge_sources (filename, file_type, record_count, status)
            VALUES (%s, %s, %s, %s)""",
            (filename, file_type, len(cleaned_data), "completed"),
        )

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "filename": filename, "file_type": file_type, "record_count": len(cleaned_data)}

    except Exception as e:
        return {"success": False, "error": str(e)}


def process_excel_legacy(file_path: str, filename: str) -> Dict[str, Any]:
    """Process Excel file - multiple sheets (legacy method)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        record_count = 0
        wb = load_workbook(file_path, read_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = [str(cell.value) if cell.value else f"col_{i}" for i, cell in enumerate(sheet[1])]

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                record = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value:
                        record[headers[i]] = str(value)

                if record:
                    cur.execute(
                        """
                        INSERT INTO knowledge_raw 
                        (source_file, file_type, sheet_name, row_number, data)
                        VALUES (%s, %s, %s, %s, %s)
                    """,
                        (filename, "excel", sheet_name, row_idx, json.dumps(record)),
                    )

                    for key, value in record.items():
                        cur.execute(
                            """
                            INSERT INTO knowledge_lookup 
                            (key_name, key_value, source_file)
                            VALUES (%s, %s, %s)
                        """,
                            (key.lower(), str(value), filename),
                        )

                    text_chunk = " | ".join([f"{k}: {v}" for k, v in record.items()])
                    cur.execute(
                        """
                        INSERT INTO knowledge_vec (content, source_file, metadata)
                        VALUES (%s, %s, %s)
                    """,
                        (text_chunk, filename, json.dumps(record)),
                    )

                    record_count += 1

        wb.close()

        cur.execute(
            """
            INSERT INTO knowledge_sources (filename, file_type, record_count, status)
            VALUES (%s, %s, %s, %s)
        """,
            (filename, "excel", record_count, "completed"),
        )

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "filename": filename, "file_type": "excel", "record_count": record_count}

    except Exception as e:
        return {"success": False, "error": str(e)}


def process_csv(file_path: str, filename: str) -> Dict[str, Any]:
    """Process CSV file."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        record_count = 0

        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)

            for row_idx, row in enumerate(reader, start=1):
                if not any(row.values()):
                    continue

                cur.execute(
                    """
                    INSERT INTO knowledge_raw 
                    (source_file, file_type, row_number, data)
                    VALUES (%s, %s, %s, %s)
                """,
                    (filename, "csv", row_idx, json.dumps(row)),
                )

                for key, value in row.items():
                    if value:
                        cur.execute(
                            """
                            INSERT INTO knowledge_lookup 
                            (key_name, key_value, source_file)
                            VALUES (%s, %s, %s)
                        """,
                            (key.lower(), str(value), filename),
                        )

                text_chunk = " | ".join([f"{k}: {v}" for k, v in row.items() if v])
                cur.execute(
                    """
                    INSERT INTO knowledge_vec (content, source_file, metadata)
                    VALUES (%s, %s, %s)
                """,
                    (text_chunk, filename, json.dumps(row)),
                )

                record_count += 1

        cur.execute(
            """
            INSERT INTO knowledge_sources (filename, file_type, record_count, status)
            VALUES (%s, %s, %s, %s)
        """,
            (filename, "csv", record_count, "completed"),
        )

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "filename": filename, "file_type": "csv", "record_count": record_count}

    except Exception as e:
        return {"success": False, "error": str(e)}


def process_word(file_path: str, filename: str) -> Dict[str, Any]:
    """Process Word document - extract tables."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        record_count = 0
        doc = Document(file_path)

        for table_idx, table in enumerate(doc.tables):
            if not table.rows:
                continue

            headers = [cell.text.strip() for cell in table.rows[0].cells]

            for row_idx, row in enumerate(table.rows[1:], start=2):
                cells = [cell.text.strip() for cell in row.cells]
                record = dict(zip(headers, cells))

                if any(record.values()):
                    cur.execute(
                        """
                        INSERT INTO knowledge_raw 
                        (source_file, file_type, sheet_name, row_number, data)
                        VALUES (%s, %s, %s, %s, %s)
                    """,
                        (filename, "word", f"table_{table_idx + 1}", row_idx, json.dumps(record)),
                    )

                    for key, value in record.items():
                        if value:
                            cur.execute(
                                """
                                INSERT INTO knowledge_lookup 
                                (key_name, key_value, source_file)
                                VALUES (%s, %s, %s)
                            """,
                                (key.lower(), str(value), filename),
                            )

                    text_chunk = " | ".join([f"{k}: {v}" for k, v in record.items() if v])
                    cur.execute(
                        """
                        INSERT INTO knowledge_vec (content, source_file, metadata)
                        VALUES (%s, %s, %s)
                    """,
                        (text_chunk, filename, json.dumps(record)),
                    )

                    record_count += 1

        cur.execute(
            """
            INSERT INTO knowledge_sources (filename, file_type, record_count, status)
            VALUES (%s, %s, %s, %s)
        """,
            (filename, "word", record_count, "completed"),
        )

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "filename": filename, "file_type": "word", "record_count": record_count}

    except Exception as e:
        return {"success": False, "error": str(e)}


def get_knowledge_sources() -> List[Dict]:
    """Get all uploaded knowledge sources."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, filename, file_type, record_count, status, created_at
            FROM knowledge_sources
            ORDER BY created_at DESC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        return [
            {
                "id": row[0],
                "filename": row[1],
                "file_type": row[2],
                "record_count": row[3],
                "status": row[4],
                "created_at": row[5].isoformat() if row[5] else None,
            }
            for row in rows
        ]
    except Exception as e:
        print(f"Error getting sources: {e}")
        return []


def delete_knowledge_source(filename: str):
    """Delete a knowledge source and all its data."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM knowledge_sources WHERE filename = %s", (filename,))
        cur.execute("DELETE FROM knowledge_raw WHERE source_file = %s", (filename,))
        cur.execute("DELETE FROM knowledge_lookup WHERE source_file = %s", (filename,))
        cur.execute("DELETE FROM knowledge_vec WHERE source_file = %s", (filename,))

        conn.commit()
        cur.close()
        conn.close()

        return True
    except Exception as e:
        print(f"Error deleting source: {e}")
        return False


def search_knowledge(query: str, limit: int = 10) -> List[Dict]:
    """Search knowledge base - exact matches."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        search_term = f"%{query.lower()}%"

        cur.execute(
            """
            SELECT DISTINCT key_name, key_value, source_file
            FROM knowledge_lookup
            WHERE key_value ILIKE %s
            LIMIT %s
        """,
            (search_term, limit),
        )

        rows = cur.fetchall()
        cur.close()
        conn.close()

        return [{"key": row[0], "value": row[1], "source": row[2]} for row in rows]
    except Exception as e:
        print(f"Error searching: {e}")
        return []


def lookup_value(key: str, value: str) -> List[Dict]:
    """Exact lookup by key-value."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT data, source_file
            FROM knowledge_raw
            WHERE data->>%s = %s
            LIMIT 20
        """,
            (key, value),
        )

        rows = cur.fetchall()
        cur.close()
        conn.close()

        return [{"data": json.loads(row[0]) if isinstance(row[0], str) else row[0], "source": row[1]} for row in rows]
    except Exception as e:
        print(f"Error looking up: {e}")
        return []


def get_source_data(filename: str, limit: int = 50) -> List[Dict]:
    """Get all data from a specific source file."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT id, data, row_number
            FROM knowledge_raw
            WHERE source_file = %s
            ORDER BY row_number
            LIMIT %s
        """,
            (filename, limit),
        )

        rows = cur.fetchall()
        cur.close()
        conn.close()

        return [
            {"id": row[0], "data": json.loads(row[1]) if isinstance(row[1], str) else row[1], "row_number": row[2]}
            for row in rows
        ]
    except Exception as e:
        print(f"Error getting source data: {e}")
        return []


def safe_json_dumps(val):
    """Safely convert a value to a JSON string, avoiding double-encoding."""
    if isinstance(val, str):
        try:
            json.loads(val)  # already valid JSON string
            return val
        except (json.JSONDecodeError, ValueError):
            return json.dumps(val)
    return json.dumps(val) if val else '[]'


def add_company(data: dict) -> dict:
    """Add a company to the companies DB table (DICA format)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Handle both old format (company_name) and new DICA format (company_name_english)
        company_name = data.get("company_name_english") or data.get("company_name", "")
        reg_no = data.get("company_registration_number", "")

        # Parse dates safely
        def safe_date(val):
            if not val:
                return None
            return val if len(str(val)) >= 8 else None

        # Ensure JSONB fields are proper
        directors = data.get("directors", [])
        if isinstance(directors, str):
            directors = [{"name": n.strip(), "type": "Director"} for n in directors.split(",") if n.strip()]
        members = data.get("members", [])
        if isinstance(members, str):
            members = []
        filing_history = data.get("filing_history", [])
        if isinstance(filing_history, str):
            filing_history = []

        cur.execute("""
            INSERT INTO companies (
                company_name_english, company_name_myanmar, company_registration_number,
                registration_date, status, company_type, foreign_company, small_company,
                principal_activity, date_of_last_annual_return, previous_registration_number,
                registered_office_address, principal_place_of_business,
                directors, ultimate_holding_company_name,
                ultimate_holding_company_jurisdiction, ultimate_holding_company_registration_number,
                total_shares_issued, currency_of_share_capital, members, filing_history,
                under_corpsec_management, group_company, total_capital, consideration_amount_paid,
                source, pdf_url
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s::jsonb, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb,
                %s, %s, %s, %s, %s, %s
            )
            ON CONFLICT (company_registration_number) DO UPDATE SET
                company_name_english = EXCLUDED.company_name_english,
                company_name_myanmar = EXCLUDED.company_name_myanmar,
                status = EXCLUDED.status, company_type = EXCLUDED.company_type,
                principal_activity = EXCLUDED.principal_activity,
                registered_office_address = EXCLUDED.registered_office_address,
                principal_place_of_business = EXCLUDED.principal_place_of_business,
                directors = EXCLUDED.directors, members = EXCLUDED.members,
                filing_history = EXCLUDED.filing_history,
                ultimate_holding_company_name = EXCLUDED.ultimate_holding_company_name,
                total_shares_issued = EXCLUDED.total_shares_issued,
                currency_of_share_capital = EXCLUDED.currency_of_share_capital,
                under_corpsec_management = EXCLUDED.under_corpsec_management,
                group_company = EXCLUDED.group_company,
                total_capital = EXCLUDED.total_capital,
                consideration_amount_paid = EXCLUDED.consideration_amount_paid,
                pdf_url = EXCLUDED.pdf_url,
                updated_at = NOW()
        """, (
            company_name,
            data.get("company_name_myanmar"),
            reg_no,
            safe_date(data.get("registration_date")),
            data.get("status"),
            data.get("company_type"),
            data.get("foreign_company"),
            data.get("small_company"),
            data.get("principal_activity"),
            safe_date(data.get("date_of_last_annual_return")),
            data.get("previous_registration_number"),
            data.get("registered_office_address") or data.get("registered_office"),
            data.get("principal_place_of_business"),
            safe_json_dumps(directors),
            data.get("ultimate_holding_company_name"),
            data.get("ultimate_holding_company_jurisdiction"),
            data.get("ultimate_holding_company_registration_number"),
            data.get("total_shares_issued"),
            data.get("currency_of_share_capital"),
            safe_json_dumps(members),
            safe_json_dumps(filing_history),
            data.get("under_corpsec_management"),
            data.get("group_company"),
            data.get("total_capital"),
            data.get("consideration_amount_paid"),
            data.get("source", "manual"),
            data.get("pdf_url"),
        ))
        conn.commit()

        # Also add to knowledge_lookup for AI search
        lookup_fields = {
            "company_name": company_name,
            "company_registration_number": reg_no,
            "registered_office": data.get("registered_office_address"),
            "directors": ", ".join(d.get("name", "") for d in directors) if isinstance(directors, list) else str(directors),
            "company_type": data.get("company_type"),
            "status": data.get("status"),
        }
        for key, value in lookup_fields.items():
            if value:
                cur.execute(
                    "INSERT INTO knowledge_lookup (key_name, key_value, source_file, created_at) VALUES (%s, %s, %s, NOW())",
                    (key, str(value), f"company:{company_name}")
                )
        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "message": f"Company '{company_name}' saved"}
    except Exception as e:
        return {"success": False, "error": str(e)}


COLUMN_MAPPING = {
    # New format -> Internal format
    "company_name": "company_name",
    "company name": "company_name",
    "company name (original)": "company_name_original",
    "registration number": "company_registration_number",
    "registration_number": "company_registration_number",
    "date of incorporation": "date_of_incorporation",
    "date_of_incorporation": "date_of_incorporation",
    "registered office address": "registered_office",
    "registered_office_address": "registered_office",
    "principal place of business": "principal_place_of_business",
    "principal_place_of_business": "principal_place_of_business",
    "company type (dica)": "company_type",
    "company_type": "company_type",
    "foreign company": "foreign_company",
    "foreign_company": "foreign_company",
    "under corpsec management": "under_corpsec_management",
    "under_corpsec_management": "under_corpsec_management",
    "group company": "group_company",
    "group_company": "group_company",
    "directors": "directors",
    "total shares": "total_shares",
    "total_shares": "total_shares",
    "total capital": "total_capital",
    "total_capital": "total_capital",
    "shareholders": "shareholders",
    "currency": "currency",
    # Meeting-specific fields (for document generation)
    "meeting_date": "meeting_date",
    "meeting_location": "meeting_location",
    "agenda": "agenda",
    "director_name": "director_name",
    "individual_shareholder_1_name": "individual_shareholder_1_name",
    "individual_shareholder_2_name": "individual_shareholder_2_name",
    "individual_shareholder_3_name": "individual_shareholder_3_name",
    "individual_shareholder_4_name": "individual_shareholder_4_name",
    "corporate_shareholder_1_name": "corporate_shareholder_1_name",
    "financial_year_end_date": "financial_year_end_date",
    "pronoun": "pronoun",
    "auditor_name": "auditor_name",
    "auditor_fee": "auditor_fee",
    "next_financial_year_end_date": "next_financial_year_end_date",
}


def map_columns_to_internal(record: dict) -> dict:
    """Map columns from new format to internal format."""
    mapped = {}
    for key, value in record.items():
        key_lower = str(key).lower().strip()
        internal_key = COLUMN_MAPPING.get(key_lower, key_lower)
        mapped[internal_key] = value
    return mapped


def process_excel_with_mapping(file_path: str, filename: str) -> Dict[str, Any]:
    """Process Excel file with column mapping for new format."""
    try:
        import pandas as pd

        df = pd.read_excel(file_path)

        # Map columns
        mapped_records = []
        for _, row in df.iterrows():
            record = row.to_dict()
            mapped = map_columns_to_internal(record)
            mapped_records.append(mapped)

        # Store in knowledge base
        conn = get_db_connection()
        cur = conn.cursor()

        record_count = 0
        for idx, record in enumerate(mapped_records, start=2):
            cur.execute(
                """INSERT INTO knowledge_raw 
                (source_file, file_type, sheet_name, row_number, data)
                VALUES (%s, %s, %s, %s, %s)""",
                (filename, "excel", "Sheet1", idx, json.dumps(record)),
            )

            for key, value in record.items():
                if value and str(value) != "nan":
                    cur.execute(
                        """INSERT INTO knowledge_lookup 
                        (key_name, key_value, source_file)
                        VALUES (%s, %s, %s)""",
                        (str(key).lower(), str(value), filename),
                    )

            text_chunk = " | ".join([f"{k}: {v}" for k, v in record.items() if v and str(v) != "nan"])
            cur.execute(
                """INSERT INTO knowledge_vec (content, source_file, metadata)
                VALUES (%s, %s, %s)""",
                (text_chunk, filename, json.dumps(record)),
            )

            record_count += 1

        cur.execute(
            """INSERT INTO knowledge_sources (filename, file_type, record_count, status)
            VALUES (%s, %s, %s, %s)""",
            (filename, "excel", record_count, "completed"),
        )

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "filename": filename, "file_type": "excel", "record_count": record_count}

    except Exception as e:
        return {"success": False, "error": str(e)}
